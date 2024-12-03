import os, sys
from openpyxl import load_workbook
import pandas as pd, numpy as np
from argparse import Namespace
from datetime import datetime
import evaluate
import pandas as pd
import pickle
import re, json
from datasets import Dataset
from transformers import AutoConfig, GPT2LMHeadModel
from transformers import AutoModelForCausalLM, Trainer, TrainingArguments, DataCollatorForLanguageModeling
from datasets import load_dataset
from transformers import AutoTokenizer, pipeline, AutoModelForSequenceClassification, TrainingArguments
from trl import AutoModelForCausalLMWithValueHead, PPOConfig, PPOTrainer, RewardTrainer, RewardConfig, get_peft_config, SFTConfig
from trl import DPOTrainer, DPOConfig
from trl.trainer.ppov2_trainer import PPOv2Trainer
from tqdm import tqdm
import transformers
import torch
from peft import LoraConfig, TaskType
import random
from functools import partial
import math

# Add the ~/myUtil directory to sys.path
sys.path.append(os.path.expanduser('~/'))
from myUtils.timeUtils import TimeUtils
from myUtils.IOUtils import IOUtils
sys.path.append(os.path.expanduser('~/2024aut_comprac/'))

from KoreanNumber import num2kr

device = "cuda" if torch.cuda.is_available() else "cpu"

import random
import numpy as np
import torch

seed = 2021
deterministic = True

random.seed(seed)
np.random.seed(seed)
torch.manual_seed(seed)
torch.cuda.manual_seed_all(seed)
if deterministic:
	torch.backends.cudnn.deterministic = True
	torch.backends.cudnn.benchmark = False

        
        
#############################################################################################
#############################################################################################
#############################################################################################
#####################################data preprocess#########################################

def _num_to_str(data):
    if isinstance(data, str):
        data = [data]

    data = [num2kr.num2kr(i, 1) for i in data]    
    
    return [re.sub(r"십([^만])", r"십만\1", i) for i in data]


def preprocess(args):

    if os.path.isfile(args.name_processed_files):
        with open(args.name_processed_files, "rb") as f:
            data = pickle.load(f)
            return data
        
    files = os.listdir(args.dir_files)
    
    data = []
    for file in files:
       
        load_wb = load_workbook(args.dir_files + os.sep + file, data_only=True)    
        load_ws = load_wb[load_wb.sheetnames[0]]

        # Extract data in a table-like format (list of dictionaries)
        header = [header for header in load_ws.iter_rows(min_row=1, max_row = 1, values_only=True)][0]
        rows = [row for row in load_ws.iter_rows(min_row=2, values_only=True)]

        x = pd.DataFrame(rows)
        x.columns = header
        x = x[args.columns_need]
        x['일자'] = x['일자'].map(lambda x : datetime.strptime(x, "%Y/%m/%d") if pd.notna(x) and x != '' else None)
        x = x.dropna(subset = ['일자'])
        # List(일자)[List(가격)[int]]
        # List(carrier[List(가격)[int]])
        data.append(x.groupby("일자")['종가'].apply(list).tolist())
    
    data = [i for j in data for i in j]
    data = [i[:int(len(i)/2)] for i in data] + [i[int(len(i)/2):] for i in data]
    with open(args.name_processed_files, "wb") as f:
        pickle.dump(data, f)

    return data

def get_distribution(data):
    d = [[abs(i2 - i1) for i1, i2 in zip(item, item[1:])] for item in data]
    d = [i for j in d for i in j]
    return round(sum(d) / len(d), 2)


# pretrain/preproces.py로  preprocess된 데이터를 한 번 더 preprocess하는 모듈
def _str_to_num(list_x):
    digits = {'일': 1, '이': 2, '삼': 3, '사': 4, '오': 5, '육': 6, '칠': 7, '팔': 8, '구': 9}
    units = {'먕': 100000, '만': 10000, '천': 1000, '백': 100, '십': 10} # 먕: 10만
    if isinstance(list_x, str):
        list_x = [list_x]

    list_out = []
    for i in list_x:            
        # 중복된 숫자 및 단위 처리
        korean_str = re.sub(r'십만', "먕", i)  # 숫자 중복 제거

        units_only = re.sub(r"[일이삼사오육칠팔구]", "_", korean_str).split("_")            
        units_only = [units[j[0]] for j in units_only if j != ""]

        digits_only = re.sub(r"[^일이삼사오육칠팔구]", "_", korean_str).split("_")
        digits_only = [digits[j[-1]] for j in digits_only if j != ""]
        
        out = 0
        for i, j in zip(units_only, digits_only):
            out += i * j
        
        list_out.append(out)
    
    return list_out


#############################################################################################
#############################################################################################
#############################################################################################
#####################################prepare dataset#########################################

def set_dpo_dataset(data):

    if isinstance(data, int):
        data = [data]

    rejected_data = []
    
    # 십만의 자리수 및 만의 자리수에서 바뀌는 데이터는 적절하지 않다. 
    for example in data:
        rejected_exmample = []
        for dataIdx in range(len(example)):
            x = int(round(random.randint(10000, 100000), -4))
            rejected_exmample.append(min(example[dataIdx] + x, 200000))
        rejected_data.append(rejected_exmample)

    rejected_data = [" ".join(_num_to_str(i)) for i in rejected_data]
    data = [" ".join(_num_to_str(i)) for i in data]
    dataset = Dataset.from_dict({"prompt": [""] * len(data), "chosen": data, "rejected": rejected_data})

    dataset_train_test = dataset.train_test_split(0.2)
    dataset_train = dataset_train_test['train']
    dataset_valid_test = dataset_train_test['test'].train_test_split(0.5)
    dataset_valid = dataset_valid_test['train']
    dataset_test = dataset_valid_test['test']

    print("reward_dataset", dataset_train)

    return dataset_train, dataset_valid, dataset_test



def set_normal_dataset(data):
    data = [" ".join(_num_to_str(i)) for i in data]
    dataset_normal = Dataset.from_dict({"text": data})

    dataset_normal_train_test = dataset_normal.train_test_split(0.2)
    dataset_normal_train = dataset_normal_train_test['train']
    dataset_normal_valid_test = dataset_normal_train_test['test'].train_test_split(0.5)
    dataset_normal_valid, dataset_normal_test = dataset_normal_valid_test['train'], dataset_normal_valid_test['test']

    print("train_dataset", dataset_normal_train)

    return dataset_normal_train, dataset_normal_valid, dataset_normal_test



#############################################################################################
#############################################################################################
#############################################################################################
#####################################prepare model###########################################

def make_config(args, tokenizer):
    
    config = AutoConfig.from_pretrained(
        args.ggangtong_model_checkpoint,
        vocab_size = len(tokenizer),
        n_ctx = 1024,
        bos_token_id = tokenizer.bos_token_id,
        eos_token_id = tokenizer.eos_token_id,
        n_embd = 512,
        n_head = 8,
        n_layer = 8,
        n_positions = 1024, 
    )

    return config


#############################################################################################
#############################################################################################
#############################################################################################
###################################prepare trainer###########################################


# def set_ppo_trainer(args, dataset, model, tokenizer, config):
def set_ppo_trainer(args, datasets, model, tokenizer, randRange):

    od = args.output_final +f"_{randRange}" + os.sep + datetime.strftime(datetime.now(), "%m-%d-%H-%M-%S")
    try: os.mkdir(od)
    except: pass

    dpoConfig = DPOConfig(
        output_dir=od,               # directory to save and repository id
        num_train_epochs=10,                     # number of training epochs
        per_device_train_batch_size=512,         # batch size per device during training
        per_device_eval_batch_size=512,           # batch size for evaluation
        optim="adamw_torch",              # use fused adamw optimizer
        learning_rate=1e-7,                     # 10x higher LR than QLoRA paper
        lr_scheduler_type="cosine",             # use cosine learning rate scheduler
        save_total_limit=1,                     # limit the total amount of checkpoints
        save_strategy="epoch",            # evaluate every 1000 steps
        logging_strategy = "epoch",
        bf16=True,                              # use bfloat16 precision
        tf32=False,                              # use tf32 precision
        push_to_hub=False,                      # push model to hub
        report_to="tensorboard",                # report metrics to tensorboard
    )



    train_dataset = datasets['train']
    valid_dataset = datasets['valid']
    # train_dataset = train_dataset.map(tokenize, batched = False)
    # valid_dataset = valid_dataset.map(tokenize, batched = False)

    dpo_args = {
        "beta":  0.1,
        "loss_type": "sigmoid"
    }

    trainer = DPOTrainer(
        model,
        ref_model=None, # set to none since we use peft
        args=dpoConfig,
        train_dataset=train_dataset,
        eval_dataset=valid_dataset,
        tokenizer=tokenizer,
        max_length=512,
        max_prompt_length=64,
        beta=dpo_args["beta"],
        loss_type=dpo_args["loss_type"],
    )
    
    return trainer


def set_normal_trainer(args, datasets, model, tokenizer):

        
    accuracy = evaluate.load('accuracy')
    def metric(eval_pred, func):
        predictions, labels = eval_pred
        predictions = np.argmax(predictions, axis = -1) # (batch, sequence lenagh, hidden_state)
        filters = labels != -100

        predictions = predictions[filters]
        labels = labels[filters]
        return func.compute(predictions = predictions, references = labels)
    
    def tokenize_func(examples):
        kwargs = {"padding": "max_length", "truncation": True, "max_length": 512, "return_tensors": "pt"}
        return tokenizer(examples['text'], **kwargs)

    train_dataset = datasets['train'].map(tokenize_func, batched=True, num_proc = 4)
    valid_dataset = datasets['valid'].map(tokenize_func, batched=True, num_proc = 4)
    train_dataset = train_dataset.remove_columns("text")
    valid_dataset = valid_dataset.remove_columns("text")

    # def tokenize(sample):
    #     out = tokenizer(sample["text"])
    #     return {"input_ids": out['input_ids']}

    # train_dataset = datasets['train']
    # valid_dataset = datasets['valid']
    # train_dataset = train_dataset.map(tokenize, batched = False)
    # valid_dataset = valid_dataset.map(tokenize, batched = False)    

    od = args.output_final + f"_normal" + os.sep + datetime.strftime(datetime.now(), "%m-%d-%H-%M-%S")
    try: os.mkdir(od)
    except: pass

    trainingarguments = TrainingArguments(
        do_train = True,    
        output_dir = od,                         
        evaluation_strategy = "epoch", # necessary: change to step
        logging_strategy = "epoch",
        save_total_limit = 1,
        greater_is_better = True, # necessary: higher metric results better performance # default = True when metric_for_best_model is set
        num_train_epochs = 10,
        seed = 42,
        per_device_train_batch_size = 512,
        per_device_eval_batch_size = 512,
        # eval_accumulation_steps = 50,
        learning_rate = 1e-7,
        bf16=True,                              # use bfloat16 precision
        tf32=False,                              # use tf32 precision
        remove_unused_columns = False
    )

    with open(od+ os.sep + "trainingargs.json", "w") as f: 
        f.write(json.dumps(trainingarguments.to_dict(), indent = 2, ensure_ascii = False))
    f.close()
    

    trainer = Trainer(
        model = model,
        args = trainingarguments,
        tokenizer = tokenizer,
        train_dataset = train_dataset,
        data_collator = DataCollatorForLanguageModeling(tokenizer=tokenizer, mlm=False),
        compute_metrics = partial(metric, func = accuracy)
    )

    return trainer


def set_smoothing_probs(args):
    # labels
    # 
    pass

def set_smoothing_CE_trainer(args, datasets, model, tokenizer):
        
    accuracy = evaluate.load('accuracy')
    def metric(eval_pred, func):
        predictions, labels = eval_pred
        predictions = np.argmax(predictions, axis = -1) # (batch, sequence lenagh, hidden_state)
        filters = labels != -100

        predictions = predictions[filters]
        labels = labels[filters]
        return func.compute(predictions = predictions, references = labels)
    
    def tokenize_func(examples):
        kwargs = {"padding": "max_length", "truncation": True, "max_length": 512, "return_tensors": "pt"}
        return tokenizer(examples['text'], **kwargs)

    train_dataset = datasets['train'].map(tokenize_func, batched=True, num_proc = 4)
    valid_dataset = datasets['valid'].map(tokenize_func, batched=True, num_proc = 4)
    train_dataset = train_dataset.remove_columns("text")
    valid_dataset = valid_dataset.remove_columns("text")

    # def tokenize(sample):
    #     out = tokenizer(sample["text"])
    #     return {"input_ids": out['input_ids']}

    # train_dataset = datasets['train']
    # valid_dataset = datasets['valid']
    # train_dataset = train_dataset.map(tokenize, batched = False)
    # valid_dataset = valid_dataset.map(tokenize, batched = False)    

    od = args.output_normal_checkpoint + os.sep + datetime.strftime(datetime.now(), "%m-%d-%H-%M-%S")
    try: os.mkdir(od)
    except: pass

    trainingarguments = TrainingArguments(
        do_train = True,    
        output_dir = od,                         
        evaluation_strategy = "steps", # necessary: change to step
        save_strategy = "steps",                         
        eval_steps = 20, # necessary: set step
        save_steps = 20,
        save_total_limit = 2,
        load_best_model_at_end = True, # necessary: EarlyStoppingCallBack하려면 True여야 함
        metric_for_best_model = "accuracy",
        greater_is_better = True, # necessary: higher metric results better performance # default = True when metric_for_best_model is set
        num_train_epochs = 10,
        seed = 42,
        per_device_train_batch_size = 512,
        per_device_eval_batch_size = 512,
        # eval_accumulation_steps = 50,
        learning_rate = 1e-7,
        remove_unused_columns = False
    )

    with open(od+ os.sep + "trainingargs.json", "w") as f: 
        f.write(json.dumps(trainingarguments.to_dict(), indent = 2, ensure_ascii = False))
    f.close()

    class SmoothingCETrainer(Trainer):
        def compute_loss(self, model, inputs, return_outputs=False):
            labels = inputs.get("labels")
            outputs = model(**inputs)
            # might be [Batch, seq_len, vocab_size], [Batch, seq_len]
            print("size", outputs.logits.size(), labels.size())

            # logs = torch.log(output.logits[-1])
            # maxv = torch.ceil(logs.max())
            # logits = maxv - logs # larger 


            logits = outputs.logits.squeeze(-1)  # Ensure logits and labels have the same shape
            mse_loss = MSELoss()(logits, labels.float())  # Convert labels to float for regression
            return (mse_loss, outputs) if return_outputs else mse_loss

    trainer = SmoothingCETrainer(
        model = model,
        args = trainingarguments,
        tokenizer = tokenizer,
        train_dataset = train_dataset,
        eval_dataset = valid_dataset,
        data_collator = DataCollatorForLanguageModeling(tokenizer=tokenizer, mlm=False),
        compute_metrics = partial(metric, func = accuracy)
    )

    return trainer



#############################################################################################
#############################################################################################
#############################################################################################
########################################decode###############################################


def generate_decode(args, model_checkpoint, dataset_test):
    model = AutoModelForCausalLM.from_pretrained(model_checkpoint)
    tokenizer = AutoTokenizer.from_pretrained(model_checkpoint)

    def tokenize_func(examples):
        return tokenizer([" ".join(i.split(" ")[:10]) for i in examples['text']], truncation=True, padding=True)

    def get_gold(examples):
        return {"gold": [" ".join(i.split(" ")[10:]) for i in examples['text']]}
    
    gold_value = dataset_test.map(get_gold, batched = True, num_proc = 4)['gold']
    test_data = dataset_test.map(tokenize_func, batched=True, num_proc = 4)

    # Generate predictions
    decoded_outputs = []
    model.eval()  # Set model to evaluation mode
    with torch.no_grad():  # Disable gradient calculation for efficiency
        for example in test_data:
            input_ids = torch.tensor(example['input_ids']).unsqueeze(0)  # Add batch dimension
            outputs = model(input_ids)
            predictions = torch.argmax(outputs.logits, dim=-1)
            
            # Decode the predicted tokens to text
            decoded_text = tokenizer.decode(predictions[0], skip_special_tokens=True)
            decoded_outputs.append(decoded_text)
    
    return decoded_outputs, gold_value

def calculate_gap(args, decoded_outputs, gold_value):

    decoded_outputs = [i.split(" ") for i in decoded_outputs]
    gold_value = [i.split(" ") for i in gold_value]

    decoded_outputs = [_str_to_num(i) for i in decoded_outputs] # List[List[int]]
    gold_value = [_str_to_num(i) for i in gold_value] # List[List[int]]

    min_length = [min(len(i), len(j)) for i, j in zip(decoded_outputs, gold_value)]

    decoded_outputs = [v[:i] for i, v in zip(min_length, decoded_outputs)]
    gold_value = [v[:i] for i, v in zip(min_length, gold_value)]
    
    res = [sum([abs(ii-jj) for ii, jj in zip(i, j)]) for i, j in zip(decoded_outputs, gold_value)]

    return sum(res) / len(res)



#############################################################################################
#############################################################################################
#############################################################################################
#########################################main################################################


@TimeUtils.consumedTime_decorator # the arguments should only be a single namespace object
def main(args):
    
    print("loading normal dataset")
    with open(args.dataset_normal, "rb") as f:
        dataset_normal = pickle.load(f)
    
    print("loading rand100 dataset")
    with open(args.dataset_rand100, "rb") as f:
        dataset_rand100 = pickle.load(f)
    
    print("loading rand1000 dataset")
    with open(args.dataset_rand1000, "rb") as f:
        dataset_rand1000 = pickle.load(f)
    
    print("loading rand10000 dataset")
    with open(args.dataset_rand10000, "rb") as f:
        dataset_rand10000 = pickle.load(f)
    
    print("prepare models")
    tokenizer = AutoTokenizer.from_pretrained(args.tokenizer_joint)
    tokenizer.add_special_tokens({"pad_token": "<pad>"}) # Llama3 doesn't have pad_token
    tokenizer.bos_token = tokenizer.eos_token
    tokenizer.padding_side = "left"
    tokenizer.truncation_side = "left"

    config = make_config(args, tokenizer)
    model = GPT2LMHeadModel(config)
    print(format(sum(p.numel() for p in model.parameters() if p.requires_grad), ",d"))

    if args.DP:
        model = DataParallel(model)
    if args.DDP:
        model.to(args.local_rank)
        model = dist(model, device_ids=[args.local_rank])
        

    print("set trainer")
    ppoModel, ppoTokenizer = model, tokenizer
    
    normalTrainer = set_normal_trainer(args, dataset_normal, model, tokenizer)
    ppoTrainer = set_ppo_trainer(args, dataset_rand100, ppoModel, ppoTokenizer, "100")
    
    ppoTrainer.train()
    normalTrainer.train()

    

    # evaluate
    #10000
    # best_rlhf = "/home/hyohyeongjang/2024aut_comprac/weights/no_rlhf_result/11-20-21-34-00/checkpoint-20"
    # best_norlhf = "/home/hyohyeongjang/2024aut_comprac/weights/rlhf_result/11-20-21-34-13/checkpoint-200"
    
    #1000
    # best_rlhf = "/home/hyohyeongjang/2024aut_comprac/weights/no_rlhf_result/11-21-11-35-39/checkpoint-20"
    # best_norlhf = "/home/hyohyeongjang/2024aut_comprac/weights/rlhf_result/11-21-11-35-52/checkpoint-200"
    
    #100
    # best_rlhf = "/home/hyohyeongjang/2024aut_comprac/weights/no_rlhf_result/11-21-11-51-41/checkpoint-20"
    # best_norlhf = "/home/hyohyeongjang/2024aut_comprac/weights/rlhf_result/11-21-11-51-54/checkpoint-200"

    # output_rlhf, gold_rlfh = generate_decode(args, best_rlhf, dataset_normal_test)
    # output_normal, gold_normal = generate_decode(args, best_norlhf, dataset_normal_test)

    # print("RLHF gap: ", calculate_gap(args, output_rlhf, gold_rlfh))
    # print("NORMAL gap: ", calculate_gap(args, output_normal, gold_normal))


if __name__ == "__main__":
    
    DDP = False
    DP = False
    global_rank = '0'
    if DP or DDP:
        os.environ["CUDA_VISIBLE_DEVICES"] = global_rank
    else:
        os.environ["CUDA_VISIBLE_DEVICES"] = '0'

    if DDP:
        import torch.distributed as dist
        dist.init_process_group("nccl")
        local_rank = int(os.environ["LOCAL_RANK"])
        torch.cuda.set_device(local_rank)
    else:
        local_rank = None

    if DP:
        from torch.nn import DataParallel

    args = Namespace(
        
        ggangtong_model_checkpoint = "openai-community/gpt2",  

        # dir_files = "/home/hyohyeongjang/2024aut_comprac/data/data_rlhf"
        # name_processed_files = "/home/hyohyeongjang/2024aut_comprac/data/data_rlhf_processed/data_rlhf.pk"
        # columns_need = ['일자', '종가']
        # checkpoint = "/home/hyohyeongjang/2024aut_comprac/weights/model_joint/10-15-02-51-17/checkpoint-50"
        tokenizer_joint = "/home/hyohyeongjang/2024aut_comprac/tokenizers/tokenizer_joint_jaeyoon",
        # reward_model_checkpoint = "distilroberta-base"
        
        output_reward_checkpoint = "/home/hyohyeongjang/2024aut_comprac/weights/reward_model/checkpoint-936",
        output_final = "/home/hyohyeongjang/2024aut_comprac/weights/final_result",
        num_cores = 4,

        dataset_normal = "/home/hyohyeongjang/2024aut_comprac/data/data_final/data-normal.pk",
        dataset_rand100 = "/home/hyohyeongjang/2024aut_comprac/data/data_final/data-reward-rand100.pk",
        dataset_rand1000 = "/home/hyohyeongjang/2024aut_comprac/data/data_final/data-reward-rand1000.pk",
        dataset_rand10000 = "/home/hyohyeongjang/2024aut_comprac/data/data_final/data-reward-rand10000.pk",

    )

    args.local_rank = local_rank
    args.DP = DP
    args.DDP = DDP

    main(args)