import os, sys
from openpyxl import load_workbook
import pandas as pd, numpy as np
from argparse import Namespace
from datetime import datetime
import evaluate
import pandas as pd
import pickle
import re, json
from datasets import Dataset
from transformers import AutoModelForCausalLM, Trainer, TrainingArguments, DataCollatorForLanguageModeling, GPT2LMHeadModel
from datasets import load_dataset
from transformers import AutoTokenizer, pipeline, AutoModelForSequenceClassification, TrainingArguments
from trl import AutoModelForCausalLMWithValueHead, PPOConfig, PPOTrainer, RewardTrainer, RewardConfig, get_peft_config, SFTConfig
from trl import DPOTrainer, DPOConfig
from trl.trainer.ppov2_trainer import PPOv2Trainer
from tqdm import tqdm
import transformers
import torch
from peft import LoraConfig, TaskType
import random
from functools import partial
import math

# Add the ~/myUtil directory to sys.path
sys.path.append(os.path.expanduser('~/'))
from myUtils.timeUtils import TimeUtils
from myUtils.IOUtils import IOUtils
from KoreanNumber import num2kr

device = "cuda" if torch.cuda.is_available() else "cpu"

import random
import numpy as np
import torch

seed = 2021
deterministic = True

random.seed(seed)
np.random.seed(seed)
torch.manual_seed(seed)
torch.cuda.manual_seed_all(seed)
if deterministic:
	torch.backends.cudnn.deterministic = True
	torch.backends.cudnn.benchmark = False

# class MyPPOTrainer(PPOTrainer):
#     def prepare_dataloader(self, dataset, data_collator=None):
#         """
#         Prepare the dataloader for training.

#         Args:
#             dataset (Union[`torch.utils.data.Dataset`, `datasets.Dataset`]):
#                 PyTorch dataset or Hugging Face dataset. If a Hugging Face dataset is passed, the dataset
#                 will be preprocessed by removing the columns that are not used by the model.
#             data_collator (Optional[function]):
#                 Data collator function.

#         Returns:
#             `torch.utils.data.DataLoader`: PyTorch dataloader
#         """
#         # if isinstance(dataset, Dataset):
#         #     dataset = self._remove_unused_columns(dataset)
#         print("----------------", dataset)
#         dataloader = torch.utils.data.DataLoader(
#             dataset,
#             batch_size=self.config.batch_size,
#             collate_fn=data_collator,
#             shuffle=True,
#             drop_last=True,
#         )
        

#         return dataloader
        
def _num_to_str(data):
    if isinstance(data, str):
        data = [data]

    data = [num2kr.num2kr(i, 1) for i in data]    
    
    return [re.sub(r"십([^만])", r"십만\1", i) for i in data]


def preprocess(args):

    if os.path.isfile(args.name_processed_files):
        with open(args.name_processed_files, "rb") as f:
            data = pickle.load(f)
            return data
        
    files = os.listdir(args.dir_files)
    
    data = []
    for file in files:
       
        load_wb = load_workbook(args.dir_files + os.sep + file, data_only=True)    
        load_ws = load_wb[load_wb.sheetnames[0]]

        # Extract data in a table-like format (list of dictionaries)
        header = [header for header in load_ws.iter_rows(min_row=1, max_row = 1, values_only=True)][0]
        rows = [row for row in load_ws.iter_rows(min_row=2, values_only=True)]

        x = pd.DataFrame(rows)
        x.columns = header
        x = x[args.columns_need]
        x['일자'] = x['일자'].map(lambda x : datetime.strptime(x, "%Y/%m/%d") if pd.notna(x) and x != '' else None)
        x = x.dropna(subset = ['일자'])
        # List(일자)[List(가격)[int]]
        # List(carrier[List(가격)[int]])
        data.append(x.groupby("일자")['종가'].apply(list).tolist())
    
    data = [i for j in data for i in j]
    data = [i[:int(len(i)/2)] for i in data] + [i[int(len(i)/2):] for i in data]
    with open(args.name_processed_files, "wb") as f:
        pickle.dump(data, f)

    return data

def get_distribution(data):
    d = [[abs(i2 - i1) for i1, i2 in zip(item, item[1:])] for item in data]
    d = [i for j in d for i in j]
    return round(sum(d) / len(d), 2)


# pretrain/preproces.py로  preprocess된 데이터를 한 번 더 preprocess하는 모듈
def _str_to_num(list_x):
    digits = {'일': 1, '이': 2, '삼': 3, '사': 4, '오': 5, '육': 6, '칠': 7, '팔': 8, '구': 9}
    units = {'먕': 100000, '만': 10000, '천': 1000, '백': 100, '십': 10} # 먕: 10만
    if isinstance(list_x, str):
        list_x = [list_x]

    list_out = []
    for i in list_x:            
        # 중복된 숫자 및 단위 처리
        korean_str = re.sub(r'십만', "먕", i)  # 숫자 중복 제거

        units_only = re.sub(r"[일이삼사오육칠팔구]", "_", korean_str).split("_")            
        units_only = [units[j[0]] for j in units_only if j != ""]

        digits_only = re.sub(r"[^일이삼사오육칠팔구]", "_", korean_str).split("_")
        digits_only = [digits[j[-1]] for j in digits_only if j != ""]
        
        out = 0
        for i, j in zip(units_only, digits_only):
            out += i * j
        
        list_out.append(out)
    
    return list_out


#############################################################################################
#############################################################################################
#############################################################################################

def set_reward_dataset(data):

    if isinstance(data, int):
        data = [data]

    rejected_data = []
    
    # 십만의 자리수 및 만의 자리수에서 바뀌는 데이터는 적절하지 않다. 
    for example in data:
        rejected_exmample = []
        for dataIdx in range(len(example)):
            x = int(round(random.randint(100, 1000), -2))
            rejected_exmample.append(min(example[dataIdx] + x, 200000))
        rejected_data.append(rejected_exmample)

    rejected_data = [" ".join(_num_to_str(i)) for i in rejected_data]
    data = [" ".join(_num_to_str(i)) for i in data]
    dataset = Dataset.from_dict({"prompt": [""] * len(data), "chosen": data, "rejected": rejected_data})

    dataset_reward_train_test = dataset.train_test_split(0.2)
    dataset_reward_train = dataset_reward_train_test['train']
    dataset_reward_valid_test = dataset_reward_train_test['test'].train_test_split(0.5)
    dataset_reward_valid = dataset_reward_valid_test['train']
    dataset_reward_test = dataset_reward_valid_test['test']

    print("reward_dataset", dataset_reward_train)

    return dataset_reward_train, dataset_reward_valid, dataset_reward_test

# def set_ppo_dataset(data):

#     data = [" ".join(_num_to_str(i)) for i in data]

#     dataset_ppo = Dataset.from_dict({"query": data})
#     dataset_ppo_train_test = dataset_ppo.train_test_split(0.2)
#     dataset_ppo_train = dataset_ppo_train_test['train']
#     dataset_ppo_valid_test = dataset_ppo_train_test['test'].train_test_split(0.5)
#     dataset_ppo_valid, dataset_ppo_test = dataset_ppo_valid_test['train'], dataset_ppo_valid_test['test']

#     return dataset_ppo_train, dataset_ppo_valid, dataset_ppo_test

def set_normal_dataset(data):
    data = [" ".join(_num_to_str(i)) for i in data]
    dataset_normal = Dataset.from_dict({"text": data})

    dataset_normal_train_test = dataset_normal.train_test_split(0.2)
    dataset_normal_train = dataset_normal_train_test['train']
    dataset_normal_valid_test = dataset_normal_train_test['test'].train_test_split(0.5)
    dataset_normal_valid, dataset_normal_test = dataset_normal_valid_test['train'], dataset_normal_valid_test['test']

    print("train_dataset", dataset_normal_train)

    return dataset_normal_train, dataset_normal_valid, dataset_normal_test



#############################################################################################
#############################################################################################
#############################################################################################


# def set_reward_model(args):
#     #Select a base model whch we need to train for reward modeling.
#     model_name = args.reward_checkpoint
#     reward_model = AutoModelForSequenceClassification.from_pretrained(model_name, num_labels=1)
#     tokenizer = AutoTokenizer.from_pretrained(model_name)
#     if tokenizer.pad_token is None:
#         tokenizer.pad_token = tokenizer.eos_token
#     reward_model.config.pad_token_id = reward_model.config.eos_token_id

#     return reward_model, tokenizer

def set_ppo_model(args):
    config = PPOConfig(
        model_name=args.checkpoint,
        learning_rate=1e-7,
    )

    model = AutoModelForCausalLMWithValueHead.from_pretrained(config.model_name)
    tokenizer = AutoTokenizer.from_pretrained(config.model_name, padding_side = "left")
    tokenizer.pad_token = tokenizer.eos_token
    model.pretrained_model.resize_token_embeddings(len(tokenizer))

    return model, tokenizer, config


def set_normal_model(args):

    model = GPT2LMHeadModel.from_pretrained(args.checkpoint, device_map = "auto")
    tokenizer = AutoTokenizer.from_pretrained(args.tokenizer_joint, padding_side = "left")    
    tokenizer.pad_token = tokenizer.eos_token
    tokenizer.padding_side = "left"
    tokenizer.truncation_side = "left"
    model.resize_token_embeddings(len(tokenizer))
    
    return model, tokenizer


#############################################################################################
#############################################################################################
#############################################################################################



# def set_reward_model_trainer(args, dataset_train, dataset_eval, model, tokenizer):

#     def formatting_func(examples):
#         kwargs = {"padding": "max_length", "truncation": True, "max_length": 512, "return_tensors": "pt"}
#         tokens_chosen = tokenizer.batch_encode_plus(examples["chosen"], **kwargs)
#         tokens_rejected = tokenizer.batch_encode_plus(examples["rejected"], **kwargs)
#         return {
#             "input_ids_chosen": tokens_chosen["input_ids"], "attention_mask_chosen": tokens_chosen["attention_mask"],
#             "input_ids_rejected": tokens_rejected["input_ids"], "attention_mask_rejected": tokens_rejected["attention_mask"]
#         }

#     formatted_dataset = {}
#     formatted_dataset['train'] = dataset_train.map(formatting_func, batched = True, num_proc = args.num_cores)
#     formatted_dataset['test'] = dataset_eval.map(formatting_func, batched = True, num_proc = args.num_cores)

#     # Configuring the training arguments
#     training_args = RewardConfig(
#         output_dir=args.output_reward_checkpoint,
#         per_device_train_batch_size=16,
#         evaluation_strategy="epoch",
#         logging_steps=1,
#         num_train_epochs = 3,
#         report_to=None,
#         # center_rewards_coefficient=0.01,
#     )

#     reward_trainer = RewardTrainer(
#         model=model,
#         tokenizer=tokenizer,
#         args=training_args,
#         train_dataset=formatted_dataset['train'],
#         eval_dataset=formatted_dataset['test'],

#     )

#     return reward_trainer


# def set_ppo_trainer(args, dataset, model, tokenizer, config):
def set_ppo_trainer(args, datasets, model, tokenizer):

    od = args.output_rlhf_checkpoint + os.sep + datetime.strftime(datetime.now(), "%m-%d-%H-%M-%S")
    try: os.mkdir(od)
    except: pass

    dpoConfig = DPOConfig(
        output_dir=od,               # directory to save and repository id
        num_train_epochs=10,                     # number of training epochs
        per_device_train_batch_size=512,         # batch size per device during training
        per_device_eval_batch_size=512,           # batch size for evaluation
        gradient_accumulation_steps=1,          # number of steps before performing a backward/update pass
        gradient_checkpointing=True,            # use gradient checkpointing to save memory
        optim="adamw_torch_fused",              # use fused adamw optimizer
        learning_rate=1e-7,                     # 10x higher LR than QLoRA paper
        max_grad_norm=0.3,                      # max gradient norm based on QLoRA paper
        lr_scheduler_type="cosine",             # use cosine learning rate scheduler
        logging_steps=25,                       # log every 25 steps
        eval_steps = 20, # necessary: set step
        save_steps = 20,                        # when to save checkpoint
        save_total_limit=2,                     # limit the total amount of checkpoints
        evaluation_strategy="steps",            # evaluate every 1000 steps
        bf16=True,                              # use bfloat16 precision
        tf32=False,                              # use tf32 precision
        push_to_hub=False,                      # push model to hub
        report_to="tensorboard",                # report metrics to tensorboard
    )



    train_dataset = datasets['train']
    valid_dataset = datasets['valid']
    # train_dataset = train_dataset.map(tokenize, batched = False)
    # valid_dataset = valid_dataset.map(tokenize, batched = False)

    dpo_args = {
        "beta":  0.1,
        "loss_type": "sigmoid"
    }

    trainer = DPOTrainer(
        model,
        ref_model=None, # set to none since we use peft
        args=dpoConfig,
        train_dataset=train_dataset,
        eval_dataset=valid_dataset,
        tokenizer=tokenizer,
        max_length=512,
        max_prompt_length=64,
        beta=dpo_args["beta"],
        loss_type=dpo_args["loss_type"],
    )
    
    return trainer


def set_normal_trainer(args, datasets, model, tokenizer):

        
    accuracy = evaluate.load('accuracy')
    def metric(eval_pred, func):
        predictions, labels = eval_pred
        predictions = np.argmax(predictions, axis = -1) # (batch, sequence lenagh, hidden_state)
        filters = labels != -100

        predictions = predictions[filters]
        labels = labels[filters]
        return func.compute(predictions = predictions, references = labels)
    
    def tokenize_func(examples):
        kwargs = {"padding": "max_length", "truncation": True, "max_length": 512, "return_tensors": "pt"}
        return tokenizer(examples['text'], **kwargs)

    train_dataset = datasets['train'].map(tokenize_func, batched=True, num_proc = 4)
    valid_dataset = datasets['valid'].map(tokenize_func, batched=True, num_proc = 4)
    train_dataset = train_dataset.remove_columns("text")
    valid_dataset = valid_dataset.remove_columns("text")

    # def tokenize(sample):
    #     out = tokenizer(sample["text"])
    #     return {"input_ids": out['input_ids']}

    # train_dataset = datasets['train']
    # valid_dataset = datasets['valid']
    # train_dataset = train_dataset.map(tokenize, batched = False)
    # valid_dataset = valid_dataset.map(tokenize, batched = False)    

    od = args.output_normal_checkpoint + os.sep + datetime.strftime(datetime.now(), "%m-%d-%H-%M-%S")
    try: os.mkdir(od)
    except: pass

    trainingarguments = TrainingArguments(
        do_train = True,    
        output_dir = od,                         
        evaluation_strategy = "steps", # necessary: change to step
        save_strategy = "steps",                         
        eval_steps = 20, # necessary: set step
        save_steps = 20,
        save_total_limit = 2,
        load_best_model_at_end = True, # necessary: EarlyStoppingCallBack하려면 True여야 함
        metric_for_best_model = "accuracy",
        greater_is_better = True, # necessary: higher metric results better performance # default = True when metric_for_best_model is set
        num_train_epochs = 10,
        seed = 42,
        per_device_train_batch_size = 512,
        per_device_eval_batch_size = 512,
        # eval_accumulation_steps = 50,
        learning_rate = 1e-7,
        remove_unused_columns = False
    )

    with open(od+ os.sep + "trainingargs.json", "w") as f: 
        f.write(json.dumps(trainingarguments.to_dict(), indent = 2, ensure_ascii = False))
    f.close()
    

    trainer = Trainer(
        model = model,
        args = trainingarguments,
        tokenizer = tokenizer,
        train_dataset = train_dataset,
        eval_dataset = valid_dataset,
        data_collator = DataCollatorForLanguageModeling(tokenizer=tokenizer, mlm=False),
        compute_metrics = partial(metric, func = accuracy)
    )

    return trainer


#############################################################################################
#############################################################################################
#############################################################################################


def ppo_trainer_train(ppoTrainer, dataset_train, tokenizer, reward_model, reward_tokenizer):
    generation_kwargs = {
        "min_length": -1,
        "max_new_tokens": 256,
        "top_k": 0.0,
        "top_p": 1.0,
        "do_sample": True,
        "pad_token_id": tokenizer.eos_token_id,
    }
    epochs = 10
    batch_size = 8

    for _ in tqdm(range(epochs), "epoch: "):
        # for batch in tqdm(ppoTrainer.prepare_dataloader(dataset_train)):
        for i in range(0, len(dataset_train), batch_size):
            batch = dataset_train[i:i + batch_size]

            query_tensors = [torch.tensor(i).to(device) for i in batch["input_ids"]]
            # query_mask = [i['attention_mask'] for i in batch]
            # query_tensors = torch.stack(query_tensors).to(device)
            # print(query_tensors.size())

            #### Get response from SFTModel
            response_tensors = ppoTrainer.generate(query_tensors, **generation_kwargs)
            batch["response"] = [tokenizer.decode(r.squeeze()) for r in response_tensors]

            #### Compute reward score
            texts = [q + r for q, r in zip(batch["query"], batch["response"])]
            texts = reward_tokenizer(texts, padding = "max_length", truncation = True, max_length = 512, return_tensors = "pt").to(device)
            pipe_outputs = reward_model(**texts)
            print(pipe_outputs)
            rewards = pipe_outputs.logits.view(-1)
            # rewards = [torch.tensor(output[1]["score"]) for output in pipe_outputs]
            # rewards = [torch.tensor(output[1]["score"]) for output in pipe_outputs]

            #### Run PPO step
            stats = ppoTrainer.step(query_tensors, response_tensors, rewards)
            ppoTrainer.log_stats(stats, batch, rewards)

    #### Save model
    ppoTrainer.save_pretrained(args.output_rlhf_checkpoint)


#############################################################################################
#############################################################################################
#############################################################################################


def generate_decode(args, model_checkpoint, dataset_test):
    model = AutoModelForCausalLM.from_pretrained(model_checkpoint)
    tokenizer = AutoTokenizer.from_pretrained(model_checkpoint)

    def tokenize_func(examples):
        return tokenizer([" ".join(i.split(" ")[:10]) for i in examples['text']], truncation=True, padding=True)

    def get_gold(examples):
        return {"gold": [" ".join(i.split(" ")[10:]) for i in examples['text']]}
    
    gold_value = dataset_test.map(get_gold, batched = True, num_proc = 4)['gold']
    test_data = dataset_test.map(tokenize_func, batched=True, num_proc = 4)

    # Generate predictions
    decoded_outputs = []
    model.eval()  # Set model to evaluation mode
    with torch.no_grad():  # Disable gradient calculation for efficiency
        for example in test_data:
            input_ids = torch.tensor(example['input_ids']).unsqueeze(0)  # Add batch dimension
            outputs = model(input_ids)
            predictions = torch.argmax(outputs.logits, dim=-1)
            
            # Decode the predicted tokens to text
            decoded_text = tokenizer.decode(predictions[0], skip_special_tokens=True)
            decoded_outputs.append(decoded_text)
    
    return decoded_outputs, gold_value

def calculate_gap(args, decoded_outputs, gold_value):

    decoded_outputs = [i.split(" ") for i in decoded_outputs]
    gold_value = [i.split(" ") for i in gold_value]

    decoded_outputs = [_str_to_num(i) for i in decoded_outputs] # List[List[int]]
    gold_value = [_str_to_num(i) for i in gold_value] # List[List[int]]

    min_length = [min(len(i), len(j)) for i, j in zip(decoded_outputs, gold_value)]

    decoded_outputs = [v[:i] for i, v in zip(min_length, decoded_outputs)]
    gold_value = [v[:i] for i, v in zip(min_length, gold_value)]
    
    res = [sum([abs(ii-jj) for ii, jj in zip(i, j)]) for i, j in zip(decoded_outputs, gold_value)]

    return sum(res) / len(res)


#############################################################################################
#############################################################################################
#############################################################################################


@TimeUtils.consumedTime_decorator # the arguments should only be a single namespace object
def main(args):
    
    random.seed(42)
    data = preprocess(args)
    print(len(data))
    random.shuffle(data)

    dist = get_distribution(data)
    x = _num_to_str(data[1234])
    decode_sample = _str_to_num(x)

    half = int(len(data)/2)
    # we don't need reward model in DPO
    # dataset_reward_train, dataset_reward_valid, dataset_reward_test = set_reward_dataset(data[:half])
    dataset_reward_train, dataset_reward_valid, dataset_reward_test = set_reward_dataset(data)
    dataset_reward = {
        "train": dataset_reward_train,
        "valid": dataset_reward_valid,
        "test": dataset_reward_test
    }
    dataset_ppo = dataset_reward
    # dataset_ppo_train, dataset_ppo_valid, dataset_ppo_test = set_ppo_dataset(data[half:])
    # dataset_ppo = {
    #     "train": dataset_ppo_train,
    #     "valid": dataset_ppo_valid,
    #     "test": dataset_ppo_test
    # }
    dataset_normal_train, dataset_normal_valid, dataset_normal_test = set_normal_dataset(data)
    dataset_normal = {
        'train': dataset_normal_train, 
        "valid": dataset_normal_valid,
        "test": dataset_normal_test
    }
    
    # normalModel, normalTokenizer = set_normal_model(args)
    # ppoModel, ppoTokenizer = normalModel, normalTokenizer
    
    # normalTrainer = set_normal_trainer(args, dataset_normal, normalModel, normalTokenizer)
    # ppoTrainer = set_ppo_trainer(args, dataset_ppo, ppoModel, ppoTokenizer)
    
    # ppoTrainer.train()
    # normalTrainer.train()

    #10000
    best_rlhf = "/home/hyohyeongjang/2024aut_comprac/weights/no_rlhf_result/11-20-21-34-00/checkpoint-20"
    best_norlhf = "/home/hyohyeongjang/2024aut_comprac/weights/rlhf_result/11-20-21-34-13/checkpoint-200"
    
    #1000
    # best_rlhf = "/home/hyohyeongjang/2024aut_comprac/weights/no_rlhf_result/11-21-11-35-39/checkpoint-20"
    # best_norlhf = "/home/hyohyeongjang/2024aut_comprac/weights/rlhf_result/11-21-11-35-52/checkpoint-200"
    
    #100
    # best_rlhf = "/home/hyohyeongjang/2024aut_comprac/weights/no_rlhf_result/11-21-11-51-41/checkpoint-20"
    # best_norlhf = "/home/hyohyeongjang/2024aut_comprac/weights/rlhf_result/11-21-11-51-54/checkpoint-200"

    output_rlhf, gold_rlfh = generate_decode(args, best_rlhf, dataset_normal_test)
    output_normal, gold_normal = generate_decode(args, best_norlhf, dataset_normal_test)

    print("RLHF gap: ", calculate_gap(args, output_rlhf, gold_rlfh))
    print("NORMAL gap: ", calculate_gap(args, output_normal, gold_normal))


if __name__ == "__main__":
    dir_files = "/home/hyohyeongjang/2024aut_comprac/data/data_rlhf"
    name_processed_files = "/home/hyohyeongjang/2024aut_comprac/data/data_rlhf_processed/data_rlhf.pk"
    columns_need = ['일자', '종가']
    checkpoint = "/home/hyohyeongjang/2024aut_comprac/weights/model_joint/10-15-02-51-17/checkpoint-50"
    tokenizer_joint = "/home/hyohyeongjang/2024aut_comprac/tokenizers/tokenizer_joint_jaeyoon"
    reward_model_checkpoint = "distilroberta-base"
    # already trained
    output_reward_checkpoint = "/home/hyohyeongjang/2024aut_comprac/weights/reward_model/checkpoint-936"
    output_rlhf_checkpoint = "/home/hyohyeongjang/2024aut_comprac/weights/rlhf_result"
    output_normal_checkpoint = "/home/hyohyeongjang/2024aut_comprac/weights/no_rlhf_result"
    num_cores = 4

    args = Namespace(
        dir_files = dir_files,
        name_processed_files = name_processed_files,
        columns_need = columns_need,
        checkpoint = checkpoint,
        tokenizer_joint = tokenizer_joint,
        reward_checkpoint = reward_model_checkpoint,
        output_reward_checkpoint = output_reward_checkpoint,
        output_rlhf_checkpoint = output_rlhf_checkpoint,
        output_normal_checkpoint = output_normal_checkpoint,
        num_cores = num_cores,
    )

    main(args)